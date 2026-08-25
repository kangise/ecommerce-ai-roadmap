"""Tenant-safe CSV evidence ingestion for marketplace agent workflows."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import re
import tempfile
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from .agents import PlatformRegistry
from .auth import AuthService
from .errors import ValidationError
from .storage import Database, Principal


@dataclass(frozen=True)
class CSVReportSpec:
    platform: str | None
    required_groups: tuple[frozenset[str], ...]


REPORT_SPECS: dict[str, CSVReportSpec] = {
    "amazon_business_report": CSVReportSpec(
        "amazon",
        (
            frozenset({"asin", "parent_asin", "child_asin"}),
            frozenset(
                {
                    "sessions",
                    "page_views",
                    "units_ordered",
                    "ordered_product_sales",
                    "unit_session_percentage",
                }
            ),
        ),
    ),
    "amazon_ads_search_term": CSVReportSpec(
        "amazon",
        (
            frozenset({"campaign_name"}),
            frozenset({"search_term"}),
            frozenset({"spend"}),
        ),
    ),
    "amazon_fba_inventory": CSVReportSpec(
        "amazon",
        (
            frozenset({"asin", "seller_sku", "sku"}),
            frozenset({"fulfillable_quantity"}),
        ),
    ),
    "amazon_returns": CSVReportSpec(
        "amazon",
        (
            frozenset({"order_id"}),
            frozenset({"return_reason"}),
        ),
    ),
    "amazon_listing": CSVReportSpec(
        "amazon",
        (
            frozenset({"asin", "seller_sku", "sku"}),
            frozenset({"title", "status"}),
        ),
    ),
    "platform_generic": CSVReportSpec(None, ()),
}


HEADER_ALIASES = {
    "asin_parent": "parent_asin",
    "asin_child": "child_asin",
    "campaign": "campaign_name",
    "customer_search_term": "search_term",
    "cost": "spend",
    "afn_fulfillable_quantity": "fulfillable_quantity",
    "available": "fulfillable_quantity",
    "amazon_order_id": "order_id",
    "reason": "return_reason",
    "item_name": "title",
}

PII_COLUMN_MARKERS = (
    "buyer_email",
    "buyer_name",
    "recipient_name",
    "shipping_address",
    "ship_address",
    "address_line",
    "phone_number",
    "buyer_phone",
)

SECRET_COLUMN_MARKERS = (
    "access_token",
    "refresh_token",
    "api_key",
    "authorization",
    "client_secret",
    "password",
    "credential",
)


@dataclass(frozen=True)
class EvidenceObjectStore:
    root: Path

    def put(self, tenant_id: str, raw: bytes) -> dict[str, Any]:
        if not re.fullmatch(r"[A-Za-z0-9-]{1,64}", tenant_id):
            raise ValidationError("tenant id is not safe for object storage")
        digest = hashlib.sha256(raw).hexdigest()
        relative = Path(tenant_id) / digest[:2] / digest
        target = self.root / relative
        target.parent.mkdir(parents=True, exist_ok=True, mode=0o700)
        if target.exists():
            if target.stat().st_size != len(raw) or hashlib.sha256(target.read_bytes()).hexdigest() != digest:
                raise ValidationError("content-addressed evidence object failed integrity validation")
        else:
            handle, temporary_name = tempfile.mkstemp(prefix=".evidence-", dir=target.parent)
            try:
                os.fchmod(handle, 0o600)
                with os.fdopen(handle, "wb") as stream:
                    stream.write(raw)
                    stream.flush()
                    os.fsync(stream.fileno())
                os.replace(temporary_name, target)
            except Exception:
                try:
                    os.unlink(temporary_name)
                except FileNotFoundError:
                    pass
                raise
        return {
            "sha256": digest,
            "byte_size": len(raw),
            "object_key": relative.as_posix(),
        }


class CSVIngestor:
    MAX_RAW_BYTES = 2_000_000
    MAX_NORMALIZED_BYTES = 800_000
    MAX_ROWS = 5_000
    MAX_COLUMNS = 200
    MAX_CELL_CHARS = 5_000

    @staticmethod
    def _normalize_header(value: str) -> str:
        normalized = unicodedata.normalize("NFKC", value).strip().lower()
        normalized = normalized.replace("%", " percentage ")
        normalized = re.sub(r"[^\w]+", "_", normalized, flags=re.UNICODE).strip("_")
        return HEADER_ALIASES.get(normalized, normalized)

    @staticmethod
    def _validate_observed_at(value: str) -> str:
        try:
            observed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        except (AttributeError, ValueError) as exc:
            raise ValidationError("observed_at must be an ISO-8601 timestamp") from exc
        if observed.tzinfo is None:
            raise ValidationError("observed_at must include a timezone")
        return value

    @staticmethod
    def _validate_filename(
        filename: str, *, allowed_suffixes: set[str] | None = None
    ) -> str:
        if not isinstance(filename, str) or not 1 <= len(filename) <= 200:
            raise ValidationError("filename must be between 1 and 200 characters")
        if (
            Path(filename).name != filename
            or "/" in filename
            or "\\" in filename
            or any(ord(char) < 32 for char in filename)
            or filename in {".", ".."}
        ):
            raise ValidationError("filename must not contain a path")
        allowed_suffixes = allowed_suffixes or {".csv", ".tsv", ".txt"}
        if Path(filename).suffix.lower() not in allowed_suffixes:
            raise ValidationError(
                "evidence import filename extension is not valid for its media type"
            )
        return filename

    @classmethod
    def parse(
        cls,
        raw: bytes,
        *,
        platform: str,
        report_type: str,
        filename: str,
        observed_at: str,
        platform_registry: PlatformRegistry,
    ) -> dict[str, Any]:
        if not raw:
            raise ValidationError("CSV upload is empty")
        if len(raw) > cls.MAX_RAW_BYTES:
            raise ValidationError("CSV upload exceeds the 2 MB limit")
        if platform not in platform_registry.entries():
            raise ValidationError(f"unsupported platform: {platform}")
        spec = REPORT_SPECS.get(report_type)
        if spec is None:
            raise ValidationError(
                "unsupported report_type; use a documented Amazon type or platform_generic"
            )
        if spec.platform is not None and platform != spec.platform:
            raise ValidationError(f"report_type {report_type} requires platform {spec.platform}")
        filename = cls._validate_filename(filename)
        observed_at = cls._validate_observed_at(observed_at)
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ValidationError("CSV must be UTF-8 encoded") from exc
        if "\x00" in text:
            raise ValidationError("CSV contains NUL bytes")
        try:
            dialect = csv.Sniffer().sniff(text[:8192], delimiters=",\t;")
        except csv.Error as exc:
            raise ValidationError("could not determine CSV delimiter") from exc
        reader = csv.DictReader(io.StringIO(text, newline=""), dialect=dialect)
        original_columns = reader.fieldnames or []
        if not 2 <= len(original_columns) <= cls.MAX_COLUMNS:
            raise ValidationError("CSV must contain between 2 and 200 columns")
        columns = [cls._normalize_header(column or "") for column in original_columns]
        if any(not column for column in columns):
            raise ValidationError("CSV contains a blank or unsupported column name")
        if len(set(columns)) != len(columns):
            raise ValidationError("CSV column names collide after normalization")
        column_mapping = {
            str(original): normalized
            for original, normalized in zip(original_columns, columns)
        }
        forbidden = [
            column
            for column in columns
            if any(marker in column for marker in (*PII_COLUMN_MARKERS, *SECRET_COLUMN_MARKERS))
        ]
        if forbidden:
            raise ValidationError(
                "CSV contains forbidden personal or secret columns: " + ", ".join(sorted(forbidden))
            )
        column_set = set(columns)
        for group in spec.required_groups:
            if not column_set.intersection(group):
                raise ValidationError(
                    f"{report_type} is missing one of the required columns: "
                    + ", ".join(sorted(group))
                )

        rows: list[dict[str, str]] = []
        blank_rows = 0
        formula_cells = 0
        for row_number, raw_row in enumerate(reader, start=2):
            extras = raw_row.get(None)
            if extras and any(str(value).strip() for value in extras):
                raise ValidationError(f"CSV row {row_number} has more fields than the header")
            normalized_row: dict[str, str] = {}
            for original, column in zip(original_columns, columns):
                value = raw_row.get(original, "")
                value = "" if value is None else str(value).strip()
                if len(value) > cls.MAX_CELL_CHARS:
                    raise ValidationError(
                        f"CSV row {row_number} column {column} exceeds 5000 characters"
                    )
                if value.startswith(("=", "+", "@")) or (
                    value.startswith("-") and len(value) > 1 and not value[1].isdigit()
                ):
                    formula_cells += 1
                normalized_row[column] = value
            if not any(normalized_row.values()):
                blank_rows += 1
                continue
            rows.append(normalized_row)
            if len(rows) > cls.MAX_ROWS:
                raise ValidationError("CSV exceeds the 5000-row limit")
        if not rows:
            raise ValidationError("CSV contains no non-empty data rows")
        serialized = json.dumps(rows, ensure_ascii=False, sort_keys=True)
        if len(serialized.encode("utf-8")) > cls.MAX_NORMALIZED_BYTES:
            raise ValidationError("normalized CSV evidence exceeds the 800 KB workflow limit")
        return {
            "platform": platform,
            "report_type": report_type,
            "filename": filename,
            "observed_at": observed_at,
            "sha256": hashlib.sha256(raw).hexdigest(),
            "delimiter": dialect.delimiter,
            "rows": rows,
            "columns": columns,
            "column_mapping": column_mapping,
            "blank_rows_skipped": blank_rows,
            "formula_cells": formula_cells,
        }


class AmazonSalesTrafficJSONFlattener:
    """Bounded normalizer for GET_SALES_AND_TRAFFIC_REPORT documents."""

    MAX_RAW_BYTES = 2_000_000
    MAX_ROWS = 5_000
    MAX_DEPTH = 20
    MAX_NODES = 100_000
    MAX_CELL_CHARS = 5_000
    COLUMNS = (
        "asin",
        "sessions",
        "units_ordered",
        "ordered_product_sales",
        "unit_session_percentage",
        "currency_code",
    )

    @classmethod
    def _bounded_json(cls, raw: bytes) -> dict[str, Any]:
        if not raw:
            raise ValidationError("Amazon sales and traffic report is empty")
        if len(raw) > cls.MAX_RAW_BYTES:
            raise ValidationError("Amazon sales and traffic report exceeds the 2 MB limit")
        try:
            value = json.loads(raw.decode("utf-8-sig"))
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise ValidationError(
                "Amazon sales and traffic report must be valid UTF-8 JSON"
            ) from exc
        nodes = 0

        def inspect(item: Any, depth: int) -> None:
            nonlocal nodes
            nodes += 1
            if nodes > cls.MAX_NODES:
                raise ValidationError("Amazon sales and traffic JSON is too complex")
            if depth > cls.MAX_DEPTH:
                raise ValidationError("Amazon sales and traffic JSON is too deeply nested")
            if isinstance(item, dict):
                for key, child in item.items():
                    if not isinstance(key, str) or len(key) > 200:
                        raise ValidationError("Amazon sales and traffic JSON has an invalid key")
                    inspect(child, depth + 1)
            elif isinstance(item, list):
                for child in item:
                    inspect(child, depth + 1)
            elif isinstance(item, str) and len(item) > cls.MAX_CELL_CHARS:
                raise ValidationError("Amazon sales and traffic JSON contains an oversized value")
            elif item is not None and not isinstance(item, (str, int, float, bool)):
                raise ValidationError("Amazon sales and traffic JSON contains an unsupported value")

        inspect(value, 0)
        if not isinstance(value, dict):
            raise ValidationError("Amazon sales and traffic report must be a JSON object")
        return value

    @classmethod
    def _cell(cls, value: Any, label: str) -> str:
        if value is None:
            return ""
        if isinstance(value, bool) or not isinstance(value, (str, int, float)):
            raise ValidationError(f"Amazon sales and traffic {label} must be scalar")
        if isinstance(value, float) and not math.isfinite(value):
            raise ValidationError(f"Amazon sales and traffic {label} must be finite")
        text = str(value).strip()
        if len(text) > cls.MAX_CELL_CHARS:
            raise ValidationError(f"Amazon sales and traffic {label} is too long")
        return text

    @classmethod
    def parse(
        cls, raw: bytes, *, filename: str, observed_at: str
    ) -> dict[str, Any]:
        filename = CSVIngestor._validate_filename(
            filename, allowed_suffixes={".json"}
        )
        observed_at = CSVIngestor._validate_observed_at(observed_at)
        document = cls._bounded_json(raw)
        entries = document.get("salesAndTrafficByAsin")
        if not isinstance(entries, list) or not entries:
            raise ValidationError(
                "Amazon sales and traffic JSON requires salesAndTrafficByAsin[]"
            )
        if len(entries) > cls.MAX_ROWS:
            raise ValidationError("Amazon sales and traffic JSON exceeds the 5000-row limit")
        rows: list[dict[str, str]] = []
        for index, entry in enumerate(entries, start=1):
            if not isinstance(entry, dict):
                raise ValidationError(
                    f"Amazon sales and traffic row {index} must be an object"
                )
            traffic = entry.get("trafficByAsin", {})
            sales = entry.get("salesByAsin", {})
            if not isinstance(traffic, dict) or not isinstance(sales, dict):
                raise ValidationError(
                    f"Amazon sales and traffic row {index} has invalid metric objects"
                )
            ordered_sales = sales.get("orderedProductSales", {})
            if ordered_sales is None:
                ordered_sales = {}
            if not isinstance(ordered_sales, dict):
                raise ValidationError(
                    f"Amazon sales and traffic row {index} has invalid orderedProductSales"
                )
            asin = entry.get("childAsin") or entry.get("parentAsin") or entry.get("asin")
            asin_text = cls._cell(asin, "asin")
            if not asin_text:
                raise ValidationError(
                    f"Amazon sales and traffic row {index} is missing an ASIN"
                )
            rows.append(
                {
                    "asin": asin_text,
                    "sessions": cls._cell(traffic.get("sessions"), "sessions"),
                    "units_ordered": cls._cell(
                        sales.get("unitsOrdered"), "unitsOrdered"
                    ),
                    "ordered_product_sales": cls._cell(
                        ordered_sales.get("amount"), "orderedProductSales.amount"
                    ),
                    "unit_session_percentage": cls._cell(
                        traffic.get("unitSessionPercentage"),
                        "unitSessionPercentage",
                    ),
                    "currency_code": cls._cell(
                        ordered_sales.get("currencyCode"),
                        "orderedProductSales.currencyCode",
                    ),
                }
            )
        serialized = json.dumps(rows, ensure_ascii=False, sort_keys=True).encode("utf-8")
        if len(serialized) > CSVIngestor.MAX_NORMALIZED_BYTES:
            raise ValidationError(
                "normalized Amazon sales and traffic evidence exceeds the 800 KB limit"
            )
        return {
            "platform": "amazon",
            "report_type": "amazon_business_report",
            "filename": filename,
            "observed_at": observed_at,
            "sha256": hashlib.sha256(raw).hexdigest(),
            "delimiter": "json",
            "rows": rows,
            "columns": list(cls.COLUMNS),
            "column_mapping": {column: column for column in cls.COLUMNS},
            "blank_rows_skipped": 0,
            "formula_cells": 0,
        }


class XLSXIngestor:
    MAX_RAW_BYTES = 5_000_000
    MAX_UNCOMPRESSED_BYTES = 50_000_000
    MAX_SHEETS = 50

    @staticmethod
    def _cell_text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (datetime, date, time)):
            return value.isoformat()
        return str(value)

    @classmethod
    def parse(
        cls,
        raw: bytes,
        *,
        platform: str,
        report_type: str,
        filename: str,
        observed_at: str,
        platform_registry: PlatformRegistry,
        sheet_name: str | None = None,
    ) -> dict[str, Any]:
        if not raw:
            raise ValidationError("XLSX upload is empty")
        if len(raw) > cls.MAX_RAW_BYTES:
            raise ValidationError("XLSX upload exceeds the 5 MB limit")
        if Path(filename).suffix.lower() != ".xlsx":
            raise ValidationError("XLSX importer accepts .xlsx files only; macros are not supported")
        CSVIngestor._validate_filename(filename, allowed_suffixes={".xlsx"})
        if not zipfile.is_zipfile(io.BytesIO(raw)):
            raise ValidationError("XLSX upload is not a valid ZIP-based workbook")
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            if any(info.flag_bits & 0x1 for info in archive.infolist()):
                raise ValidationError("encrypted XLSX files are not supported")
            if sum(info.file_size for info in archive.infolist()) > cls.MAX_UNCOMPRESSED_BYTES:
                raise ValidationError("XLSX uncompressed content exceeds the 50 MB limit")
            if any(".." in Path(info.filename).parts for info in archive.infolist()):
                raise ValidationError("XLSX archive contains an unsafe path")
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            from .errors import ConnectorNotConfiguredError

            raise ConnectorNotConfiguredError(
                "XLSX support requires the ecommerce-ai-skills[xlsx] extra"
            ) from exc
        try:
            workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=False)
        except Exception as exc:
            raise ValidationError("XLSX workbook could not be opened") from exc
        try:
            if len(workbook.sheetnames) > cls.MAX_SHEETS:
                raise ValidationError("XLSX exceeds the 50-sheet limit")
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ValidationError(f"XLSX sheet not found: {sheet_name}")
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
                sheet_name = worksheet.title
            rows_iter = worksheet.iter_rows()
            try:
                header_cells = next(rows_iter)
            except StopIteration as exc:
                raise ValidationError("XLSX sheet is empty") from exc
            headers = [cls._cell_text(cell.value).strip() for cell in header_cells]
            buffer = io.StringIO(newline="")
            writer = csv.writer(buffer)
            writer.writerow(headers)
            formula_cells = 0
            for row_number, cells in enumerate(rows_iter, start=2):
                if any(cls._cell_text(cell.value).strip() for cell in cells[len(headers) :]):
                    raise ValidationError(
                        f"XLSX row {row_number} has values beyond the header columns"
                    )
                values = []
                for cell in cells[: len(headers)]:
                    if getattr(cell, "data_type", None) == "f":
                        formula_cells += 1
                    values.append(cls._cell_text(cell.value))
                if len(values) < len(headers):
                    values.extend([""] * (len(headers) - len(values)))
                writer.writerow(values)
                if row_number > CSVIngestor.MAX_ROWS + 1:
                    raise ValidationError("XLSX exceeds the 5000-row limit")
            generated = buffer.getvalue().encode("utf-8")
            parsed = CSVIngestor.parse(
                generated,
                platform=platform,
                report_type=report_type,
                filename="normalized.csv",
                observed_at=observed_at,
                platform_registry=platform_registry,
            )
            parsed.update(
                {
                    "filename": filename,
                    "sha256": hashlib.sha256(raw).hexdigest(),
                    "delimiter": "xlsx",
                    "formula_cells": formula_cells,
                    "sheet_name": sheet_name,
                }
            )
            return parsed
        finally:
            workbook.close()


class EvidenceImportService:
    def __init__(
        self,
        db: Database,
        auth: AuthService,
        *,
        platform_registry: PlatformRegistry | None = None,
        object_store: EvidenceObjectStore | None = None,
    ):
        self.db = db
        self.auth = auth
        self.platform_registry = platform_registry or PlatformRegistry()
        self.object_store = object_store or EvidenceObjectStore(
            db.path.parent / f"{db.path.name}.evidence_objects"
        )

    def _persist(
        self,
        principal: Principal,
        parsed: dict[str, Any],
        *,
        raw: bytes,
        media_type: str,
        idempotency_key: str,
        request_id: str,
    ) -> dict[str, Any]:
        stored_object = self.object_store.put(principal.tenant_id, raw)
        if parsed["sha256"] != stored_object["sha256"]:
            raise ValidationError("parsed evidence digest does not match stored object")
        parsed = dict(parsed) | stored_object | {"media_type": media_type}
        parsed.setdefault("sheet_name", None)
        imported, replayed = self.db.create_evidence_import(
            principal.tenant_id,
            principal.user_id,
            idempotency_key,
            **parsed,
        )
        self.db.append_audit(
            principal.tenant_id,
            principal.user_id,
            request_id,
            "evidence_import.create",
            "evidence_import",
            imported["id"],
            "replayed" if replayed else "accepted",
            {
                "platform": imported["platform"],
                "report_type": imported["report_type"],
                "row_count": imported["row_count"],
                "sha256": imported["sha256"],
                "media_type": imported["media_type"],
            },
        )
        return {key: value for key, value in imported.items() if key != "rows"}

    def import_csv(
        self,
        principal: Principal,
        *,
        raw: bytes,
        platform: str,
        report_type: str,
        filename: str,
        observed_at: str,
        idempotency_key: str,
        request_id: str,
        media_type: str = "text/csv",
    ) -> dict[str, Any]:
        self.auth.require(principal, "operator")
        parsed = CSVIngestor.parse(
            raw,
            platform=platform,
            report_type=report_type,
            filename=filename,
            observed_at=observed_at,
            platform_registry=self.platform_registry,
        )
        return self._persist(
            principal,
            parsed,
            raw=raw,
            media_type=media_type,
            idempotency_key=idempotency_key,
            request_id=request_id,
        )

    def import_xlsx(
        self,
        principal: Principal,
        *,
        raw: bytes,
        platform: str,
        report_type: str,
        filename: str,
        observed_at: str,
        sheet_name: str | None,
        idempotency_key: str,
        request_id: str,
    ) -> dict[str, Any]:
        self.auth.require(principal, "operator")
        parsed = XLSXIngestor.parse(
            raw,
            platform=platform,
            report_type=report_type,
            filename=filename,
            observed_at=observed_at,
            platform_registry=self.platform_registry,
            sheet_name=sheet_name,
        )
        return self._persist(
            principal,
            parsed,
            raw=raw,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            idempotency_key=idempotency_key,
            request_id=request_id,
        )

    def import_amazon_sales_traffic_json(
        self,
        principal: Principal,
        *,
        raw: bytes,
        filename: str,
        observed_at: str,
        idempotency_key: str,
        request_id: str,
    ) -> dict[str, Any]:
        self.auth.require(principal, "operator")
        parsed = AmazonSalesTrafficJSONFlattener.parse(
            raw, filename=filename, observed_at=observed_at
        )
        return self._persist(
            principal,
            parsed,
            raw=raw,
            media_type="application/json",
            idempotency_key=idempotency_key,
            request_id=request_id,
        )

    def list(self, principal: Principal, limit: int = 100) -> list[dict[str, Any]]:
        self.auth.require(principal, "viewer")
        return self.db.list_evidence_imports(principal.tenant_id, limit)

    def get(self, principal: Principal, import_id: str) -> dict[str, Any]:
        self.auth.require(principal, "viewer")
        return self.db.get_evidence_import(principal.tenant_id, import_id)

    def resolve(self, principal: Principal, import_ids: list[str]) -> list[dict[str, Any]]:
        self.auth.require(principal, "viewer")
        if not isinstance(import_ids, list) or len(import_ids) > 20:
            raise ValidationError("evidence_import_ids must contain at most 20 IDs")
        if len(set(import_ids)) != len(import_ids):
            raise ValidationError("evidence_import_ids must not contain duplicates")
        evidence = []
        for import_id in import_ids:
            if not isinstance(import_id, str) or not import_id:
                raise ValidationError("each evidence import ID must be a non-empty string")
            imported = self.db.get_evidence_import(
                principal.tenant_id, import_id, include_rows=True
            )
            evidence.append(
                {
                    "source_id": f"evidence_import:{imported['id']}",
                    "platform": imported["platform"],
                    "source_type": imported["report_type"],
                    "observed_at": imported["observed_at"],
                    "data": {
                        "filename": imported["filename"],
                        "sha256": imported["sha256"],
                        "row_count": imported["row_count"],
                        "columns": imported["columns"],
                        "column_mapping": imported["column_mapping"],
                        "formula_cells": imported["formula_cells"],
                        "media_type": imported["media_type"],
                        "byte_size": imported["byte_size"],
                        "object_key": imported["object_key"],
                        "sheet_name": imported["sheet_name"],
                        "rows": imported["rows"],
                    },
                }
            )
        return evidence
